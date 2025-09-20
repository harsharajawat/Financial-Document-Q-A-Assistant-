import re
import os
import json
import math
from rapidfuzz import fuzz
from io import BytesIO
from typing import List, Tuple
import openpyxl
from tempfile import NamedTemporaryFile
import pdfplumber
from pdf2image import convert_from_bytes
import pytesseract
import pandas as pd
from openpyxl import load_workbook
from langchain.text_splitter import RecursiveCharacterTextSplitter
import requests
from langchain.schema import Document
from langchain_community.vectorstores import Chroma
from langchain_huggingface import HuggingFaceEmbeddings
from langchain_core.runnables import RunnablePassthrough, RunnableParallel
from langchain_core.output_parsers import StrOutputParser
from langchain_core.prompts import PromptTemplate
from langchain_experimental.text_splitter import SemanticChunker
import streamlit as st
from dotenv import load_dotenv

load_dotenv()


def clean_text(text: str) -> str:
    if not isinstance(text, str):
        return ""
    text = re.sub(r"\s+", " ", text)
    text = re.sub(r"Page\s+\d+\s+of\s+\d+", "", text, flags=re.IGNORECASE)
    return text.strip()

def normalize_column_name(col: str) -> str:
    if col is None:
        return ""
    col = str(col).strip().lower()
    col = re.sub(r"\s+", "_", col)
    col = re.sub(r"[^\w]", "", col)
    return col

def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
    if not text:
        return []
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=chunk_size,
        chunk_overlap=overlap,
        separators=["\n\n", "\n", " "],
    )
    return splitter.split_text(text)

def table_to_json(table: List[List[str]]) -> List[dict]:
    if not table or len(table) < 2:
        return []
    headers = [h if h is not None else f"col_{i}" for i, h in enumerate(table[0])]
    json_data = []
    for row in table[1:]:
        row_padded = list(row) + [None] * (len(headers) - len(row))
        row_dict = {str(headers[i]): row_padded[i] for i in range(len(headers))}
        json_data.append(row_dict)
    return json_data

# -----------------------
# Normalizer Functions
# -----------------------
def clean_number(value):
    if not isinstance(value, str):
        return value
    value = value.replace("$", "").replace(",", "").replace("(", "-").replace(")", "").strip()
    try:
        if '.' in value:
            return float(value)
        else:
            return int(value)
    except ValueError:
        return value

def normalize_financial_dict(raw_dict):
    key_map = {
        "service_revenue": "Service Revenue",
        "sales_revenue": "Sales Revenue",
        "total_revenue": "Revenue",
        "operating_expenses": "Total Operating Expenses",
        "depreciation_expense": "Depreciation Expense",
        "wages_expenses": "Wages Expense",
        "supplies_expenses": "Supplies Expense",
        "operating_income": "Operating Income",
        "interest_expense": "Interest Expense",
        "pretax_income": "Pretax Income",
        "income_tax_expense": "Income Tax Expense",
        "net_income": "Net Income",
        "retained_earnings": "Retained Earnings",
        "dividends_declared": "Dividends Declared",
        "total_assets": "Total Assets",
        "total_liabilities": "Total Liabilities",
        "total_equity": "Total Equity",
        "year": "Year",
    }
    normalized = {}
    for k, v in raw_dict.items():
        if k is None:
            continue
        key = str(k).strip().lower().replace(" ", "_")
        std_key = key_map.get(key, k.title().strip())
        normalized[std_key] = clean_number(v)
    return normalized

# -----------------------
# Extract key-value pairs from free text using regex
# -----------------------
def extract_key_value_from_text(text):
    patterns = {
        "service_revenue": r"(service\s+revenue)[^\d\$]*\$?([\d,\.]+)",
        "sales_revenue": r"(sales\s+revenue)[^\d\$]*\$?([\d,\.]+)",
        "net_income": r"(net\s+income)[^\d\$]*\$?([\d,\.]+)",
        "operating_expenses": r"(operating\s+expenses)[^\d\$]*\$?([\d,\.]+)",
        "income_tax_expense": r"(income\s+tax\s+expense)[^\d\$]*\$?([\d,\.]+)",
        "pretax_income": r"(pretax\s+income)[^\d\$]*\$?([\d,\.]+)",
    }
    extracted = {}
    for key, pattern in patterns.items():
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            extracted[key] = match.group(2)
    return extracted

# Helper function to save JSON logs
def save_json(data, filename):
    os.makedirs("logs", exist_ok=True)
    with open(os.path.join("logs", filename), "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)

# -----------------------
# PDF Extraction with normalization and print key-values
# -----------------------
def extract_pdf(file_path: str) -> Tuple[List[str], List[str], List[dict]]:
    text_chunks, table_summaries = [], []
    all_extracted_dicts = []

    with open(file_path, "rb") as file:
        raw = file.read()
        if not raw:
            return text_chunks, table_summaries, all_extracted_dicts

        with pdfplumber.open(BytesIO(raw)) as pdf:
            for page in pdf.pages:
                # Extract text and chunk it
                text = page.extract_text()
                if text:
                    text_chunks.extend(chunk_text(clean_text(text)))
                    # Extract and normalize key-values from page text
                    extracted_from_text = extract_key_value_from_text(text)
                    if extracted_from_text:
                        normalized_text = normalize_financial_dict(extracted_from_text)
                        all_extracted_dicts.append(normalized_text)

                # Extract tables
                try:
                    tables = page.extract_tables()
                except Exception:
                    tables = []

                for table in tables or []:
                    table_json = table_to_json(table)
                    if table_json:
                        # Normalize each row before appending
                        normalized_rows = [normalize_financial_dict(row) for row in table_json]
                        # Append normalized json rows per table
                        table_summaries.append(json.dumps(normalized_rows))
                        all_extracted_dicts.extend(normalized_rows)

        # OCR fallback if no text or tables found
        if not text_chunks and not table_summaries:
            try:
                images = convert_from_bytes(raw)
                for img in images:
                    ocr_text = pytesseract.image_to_string(img)
                    text_chunks.extend(chunk_text(clean_text(ocr_text)))
                    extracted_from_text = extract_key_value_from_text(ocr_text)
                    if extracted_from_text:
                        normalized_text = normalize_financial_dict(extracted_from_text)
                        all_extracted_dicts.append(normalized_text)
            except Exception:
                pass

    # Save logs
    save_json(text_chunks, "pdf_raw_text_chunks.json")
    tables_json = [json.loads(t) for t in table_summaries]
    save_json(tables_json, "pdf_raw_tables.json")
    save_json(all_extracted_dicts, "pdf_normalized_dicts.json")

    # Print extracted financial key-values
    # print("\n===== Extracted Financial Key-Value Pairs (PDF) =====")
    for item in all_extracted_dicts:
        print(item)
    print("================================================\n")

    return text_chunks, table_summaries, all_extracted_dicts

#-----------------------
# Excel Extraction with normalization
#-----------------------
def unmerge_and_propagate_merged_cells(ws):
    merged_cells = list(ws.merged_cells)
    for merged_cell_range in merged_cells:
        ws.unmerge_cells(str(merged_cell_range))
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(str(merged_cell_range))
        top_left_value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value

def extract_excel_content(file) -> Tuple[List[str], List[str], List[dict]]:
    text_chunks, table_summaries = [], []
    all_extracted_dicts = []

    try:
        file.seek(0)
    except Exception:
        pass
    raw = file.read()
    if not raw:
        return text_chunks, table_summaries, all_extracted_dicts

    in_memory_file = BytesIO(raw)
    try:
        wb = openpyxl.load_workbook(in_memory_file, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            unmerge_and_propagate_merged_cells(ws)

            tmp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmp_file.name)

            # Set header row index based on your files; adjust if needed
            header_row_idx = 3
            df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, header=header_row_idx)
            tmp_file.close()

            # Normalize column names
            df.columns = [normalize_column_name(c) for c in df.columns]

            # Drop fully empty rows and columns
            df.dropna(axis=1, how='all', inplace=True)
            df.dropna(axis=0, how='all', inplace=True)

            # Rename the first column to 'description'
            if df.columns.size > 0:
                df.rename(columns={df.columns[0]: "description"}, inplace=True)

            # Remove rows with metadata keywords in description
            meta_keywords = ['template', 'related templates', 'more ', 'income statement', 'cash flow statement', 'financial ratios']
            df = df[~df['description'].fillna('').str.lower().str.contains('|'.join(meta_keywords))]

            # Keep only rows with at least one numeric value
            numeric_cols = df.select_dtypes(include=[float, int]).columns
            if len(numeric_cols) > 0:
                df = df[df[numeric_cols].notna().any(axis=1)]

            # Extend normalized dicts and summaries
            raw_dicts = df.to_dict(orient="records")
            normalized_dicts = [normalize_financial_dict(row) for row in raw_dicts]

            table_summaries.append(json.dumps({sheet_name: normalized_dicts}))
            all_extracted_dicts.extend(normalized_dicts)

            # Add text chunk (joining all text cells) for retrieval context
            text_chunks.extend(chunk_text(" ".join(df.astype(str).agg(" ".join, axis=1))))

    except Exception:
        pass

    # Save logs for debugging
    save_json(text_chunks, "excel_raw_text_chunks.json")
    tables_json = [json.loads(t) for t in table_summaries]
    save_json(tables_json, "excel_raw_tables.json")
    save_json(all_extracted_dicts, "excel_normalized_dicts.json")

    # Print extracted normalized data for debug
    # print("\n===== Filtered and Normalized Extracted Financial Key-Value Pairs (Excel) =====")
    for item in all_extracted_dicts:
        print(item)
    print("============================================================\n")

    return text_chunks, table_summaries, all_extracted_dicts


def unmerge_and_propagate_merged_cells(ws):
    merged_cells = list(ws.merged_cells)
    for merged_cell_range in merged_cells:
        ws.unmerge_cells(str(merged_cell_range))
        min_col, min_row, max_col, max_row = openpyxl.utils.range_boundaries(str(merged_cell_range))
        top_left_value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value

def extract_excel_content(file) -> Tuple[List[str], List[str], List[dict]]:
    text_chunks, table_summaries = [], []
    all_extracted_dicts = []

    try:
        file.seek(0)
    except Exception:
        pass
    raw = file.read()
    if not raw:
        return text_chunks, table_summaries, all_extracted_dicts

    in_memory_file = BytesIO(raw)
    try:
        wb = openpyxl.load_workbook(in_memory_file, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            unmerge_and_propagate_merged_cells(ws)

            tmp_file = NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(tmp_file.name)

            # Set header row index based on your files; adjust if needed
            header_row_idx = 3
            df = pd.read_excel(tmp_file.name, sheet_name=sheet_name, header=header_row_idx)
            tmp_file.close()

            # Normalize column names
            df.columns = [normalize_column_name(c) for c in df.columns]

            # Drop fully empty rows and columns
            df.dropna(axis=1, how='all', inplace=True)
            df.dropna(axis=0, how='all', inplace=True)

            # Rename the first column to 'description'
            if df.columns.size > 0:
                df.rename(columns={df.columns[0]: "description"}, inplace=True)

            # Remove rows with metadata keywords in description
            meta_keywords = ['template', 'related templates', 'more ', 'income statement', 'cash flow statement', 'financial ratios']
            df = df[~df['description'].fillna('').str.lower().str.contains('|'.join(meta_keywords))]

            # Keep only rows with at least one numeric value
            numeric_cols = df.select_dtypes(include=[float, int]).columns
            if len(numeric_cols) > 0:
                df = df[df[numeric_cols].notna().any(axis=1)]

            # Extend normalized dicts and summaries
            raw_dicts = df.to_dict(orient="records")
            normalized_dicts = [normalize_financial_dict(row) for row in raw_dicts]

            table_summaries.append(json.dumps({sheet_name: normalized_dicts}))
            all_extracted_dicts.extend(normalized_dicts)

            # Add text chunk (joining all text cells) for retrieval context
            text_chunks.extend(chunk_text(" ".join(df.astype(str).agg(" ".join, axis=1))))

    except Exception:
        pass

    # Save logs for debugging
    save_json(text_chunks, "excel_raw_text_chunks.json")
    tables_json = [json.loads(t) for t in table_summaries]
    save_json(tables_json, "excel_raw_tables.json")
    save_json(all_extracted_dicts, "excel_normalized_dicts.json")

    # Print extracted normalized data for debug
    # print("\n===== Filtered and Normalized Extracted Financial Key-Value Pairs (Excel) =====")
    # for item in all_extracted_dicts:
    #     print(item)
    # print("============================================================\n")

    return text_chunks, table_summaries, all_extracted_dicts
# -----------------------

def find_answer_in_normalized_data_w_yrs(query: str, normalized_dicts: list) -> str:
    import re
    import math
    from rapidfuzz import fuzz, process

    def is_nan(val):
        try:
            return isinstance(val, float) and math.isnan(val)
        except:
            return False

    query_lower = query.lower()

    # Extract year if present
    year_match = re.search(r"(20\d{2})", query_lower)
    year = year_match.group(1) if year_match else None

    # Normalize query keywords relevant for matching keys
    stopwords = {"what", "is", "the", "number", "of", "in", "for", "and", "was", year or ""}
    query_words = [w for w in re.findall(r"\w+", query_lower) if w not in stopwords]
    query_joined = " ".join(query_words)

    # Collect all unique keys from all dicts, normalized to lowercase
    all_keys = set()
    for d in normalized_dicts:
        for k in d.keys():
            if k:
                all_keys.add(k.lower())

    # Try direct key matching: find keys where all their words appear in query
    potential_keys = []
    for key in all_keys:
        if all(word in query_joined for word in key.split()):
            potential_keys.append(key)

    # If no direct key match found, fallback to fuzzy matching on keys
    if not potential_keys:
        key_list = list(all_keys)
        matches = process.extract(query_joined, key_list, limit=3, score_cutoff=60)
        potential_keys = [m[0] for m in matches] if matches else []

    # If direct or fuzzy key match found, try to get a value from dicts
    if potential_keys:
        for key in potential_keys:
            for d in normalized_dicts:
                # Check year value preference if year key in dict
                if year and year in d and not is_nan(d[year]) and isinstance(d[year], (int, float)):
                    return str(d[year])
                # Return matching key value if valid number
                for k, v in d.items():
                    if k.lower() == key and v is not None and not is_nan(v) and isinstance(v, (int, float)):
                        return str(v)
        return ""

    # If no direct key match found, enable fuzzy match on description-like keys for Excel-type tables
    possible_desc_keys = ["description", "unnamed_2", "assets", "Description", "Unnamed_2", "Assets"]

    best_match_score = 0
    best_match_value = ""
    for data_dict in normalized_dicts:
        for desc_key in possible_desc_keys:
            val = data_dict.get(desc_key)
            if isinstance(val, str) and val:
                desc_str = val.strip().lower()
                score = fuzz.token_sort_ratio(query_joined, desc_str)

                if score > best_match_score and score >= 60:
                    # Prefer value for year key if present and valid
                    if year and year in data_dict:
                        year_val = data_dict.get(year)
                        if year_val is not None and not is_nan(year_val) and isinstance(year_val, (int, float)):
                            best_match_score = score
                            best_match_value = str(year_val)
                            continue

                    # Otherwise, try return any numeric year column value
                    for k, value in data_dict.items():
                        if re.match(r"20\d{2}", str(k)) and value is not None and not is_nan(value) and isinstance(value, (int, float)):
                            best_match_score = score
                            best_match_value = str(value)
                            break
    if best_match_value:
        return best_match_value

    return ""



# New combined helper to extract & query in one call
def extract_and_query(file_path_or_obj, query: str, file_type: str = None) -> str:
    """
    Extracts data from given PDF or Excel file path/bytes and answers the query.
    Parameters:
        - file_path_or_obj: str file path or file-like object (e.g. BytesIO)
        - query: question string to answer from extracted data
        - file_type: explicitly set 'pdf' or 'excel' if needed (optional)
    Returns:
        - str: extracted answer or empty string if none found
    """

    # Determine extraction method
    is_pdf = False
    is_excel = False
    if file_type:
        is_pdf = file_type.lower() == "pdf"
        is_excel = file_type.lower() == "excel"
    else:
        # Infer from filename or type
        if isinstance(file_path_or_obj, str):
            is_pdf = file_path_or_obj.lower().endswith(".pdf")
            is_excel = file_path_or_obj.lower().endswith((".xls", ".xlsx"))
        else:
            # if file-like, require file_type arg or guess from BytesIO content maybe
            raise ValueError("file_type must be specified for file-like objects")

    # Extract data
    if is_pdf:
        text_chunks, table_summaries, extract_value = extract_pdf(file_path_or_obj)
        # parse normalized dicts from saved logs or reconstruct here if needed
        # For simplicity, we'll read from the log file saved by extract_pdf
        try:
            with open("logs/pdf_normalized_dicts.json", "r", encoding="utf-8") as f:
                normalized_dicts = json.load(f)
        except Exception:
            normalized_dicts = []
    elif is_excel:
        if isinstance(file_path_or_obj, str):
            with open(file_path_or_obj, "rb") as f:
                text_chunks, table_summaries, extract_value = extract_excel_content(f)
        else:
            text_chunks, table_summaries, extract_value = extract_excel_content(file_path_or_obj)
        try:
            with open("logs/excel_normalized_dicts.json", "r", encoding="utf-8") as f:
                normalized_dicts = json.load(f)
        except Exception:
            normalized_dicts = []
    else:
        raise ValueError("Unsupported file type for extraction.")

    # Query extracted data
    answer = find_answer_in_normalized_data_w_yrs(query, normalized_dicts)
    return answer



def raw_text_to_documents(text_chunks):
    return [Document(page_content=chunk) for chunk in text_chunks if chunk.strip()]

def build_rag_pipeline(raw_text_chunks):
    embedder = HuggingFaceEmbeddings(model_name="sentence-transformers/all-mpnet-base-v2")
    docs = raw_text_to_documents(raw_text_chunks)
    text_splitter = SemanticChunker(embedder)
    splits = text_splitter.split_documents(docs)

    vectorstore = Chroma.from_documents(documents=splits, embedding=embedder)
    retriever = vectorstore.as_retriever(search_type="similarity", search_kwargs={"k": 5})

    prompt_template = """
    You are a helpful financial assistant. Use ONLY the context below to answer the question. 
    If the answer is not contained in the context, politely respond that the information is not available.

    Context:
    {context}

    Question: {question}
    Answer:"""
    prompt = PromptTemplate(input_variables=["context", "question"], template=prompt_template)

    from langchain_huggingface import HuggingFaceEndpoint, ChatHuggingFace
    llm = HuggingFaceEndpoint(
    repo_id="meta-llama/Llama-3.1-8B-Instruct",
    task="text-generation"
)
    
    model = ChatHuggingFace(llm=llm)    

    def format_docs(docs):
        return "\n\n".join(doc.page_content for doc in docs)
    
    rag_chain_from_docs = (
        RunnablePassthrough.assign(context=(lambda x: format_docs(x["context"])))
        | prompt
        | model
        | StrOutputParser()
    )

    # def retriever_for_question(input_dict):
    #     question = input_dict["question"]
    #     return retriever.invoke(question)

    rag_chain_with_source = RunnableParallel(
    {
        # retrieve context by extracting "question" string and passing to retriever.invoke
        "context": (RunnablePassthrough() | (lambda x: retriever.invoke(x["question"]))),
        "question": RunnablePassthrough(),
    }
    ).assign(answer=rag_chain_from_docs)


    return rag_chain_with_source

# # Example usage/testing block
# if __name__ == "__main__":
#     # Provide some example text chunks
#     example_chunks = [
#         "RAG (Retrieval-Augmented Generation) helps combine the power of retrieval and language models.",
#         "It improves answer quality by grounding generated text in relevant context."
#     ]
    
#     rag_pipeline = build_rag_pipeline(example_chunks)
#     test_input = {"question": "What are the advantages of RAG?"}
#     result = rag_pipeline.invoke(test_input)
#     print(result["answer"])

import math

def is_empty_answer(answer):
    if answer is None:
        return True
    if isinstance(answer, str):
        if answer.strip() == "" or answer.strip().lower() == "nan":
            return True
    try:
        if isinstance(answer, float) and math.isnan(answer):
            return True
    except:
        pass
    return False

rag_pipeline = None

def init_rag_pipeline(text_chunks):
    global rag_pipeline
    if rag_pipeline is None:
        rag_pipeline = build_rag_pipeline(text_chunks)
    return rag_pipeline

def process_uploaded_file(uploaded_file):
    """
    Calls PDF or Excel extraction functions depending on file extension.
    Returns text_chunks, table_summaries, normalized_dicts extracted from file.
    """
    if uploaded_file.name.lower().endswith(".pdf"):
        # Save temp PDF file, as extract_pdf expects a filepath
        temp_file_path = "temp_uploaded.pdf"
        with open(temp_file_path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return extract_pdf(temp_file_path)
    elif uploaded_file.name.lower().endswith((".xls", ".xlsx")):
        # Excel extraction accepts file-like object directly
        uploaded_file.seek(0)  # Reset pointer
        return extract_excel_content(uploaded_file)
    else:
        return [], [], []


def main():
    global rag_pipeline

    st.title("Financial Document Q&A Assistant")
    st.markdown("""
        Upload a financial document (PDF or Excel) containing financial statements.
        Ask questions about revenue, expenses, profits, and other metrics extracted from the file.
    """)

    uploaded_file = st.file_uploader("Upload PDF or Excel financial statement", type=["pdf", "xlsx", "xls"])

    if uploaded_file:
        try:
            if 'extracted_data' not in st.session_state or st.session_state.get('uploaded_file_name') != uploaded_file.name:
                st.info(f"Processing file: {uploaded_file.name}")
                # Extract only once and save results
                text_chunks, table_summaries, normalized_dicts = process_uploaded_file(uploaded_file)

                if not normalized_dicts:
                    st.warning("Could not extract any financial data from the document.")
                    return

                st.session_state['extracted_data'] = {
                    'text_chunks': text_chunks,
                    'table_summaries': table_summaries,
                    'normalized_dicts': normalized_dicts
                }
                st.session_state['uploaded_file_name'] = uploaded_file.name
            else:
                # Reuse the previously extracted data
                text_chunks = st.session_state['extracted_data']['text_chunks']
                table_summaries = st.session_state['extracted_data']['table_summaries']
                normalized_dicts = st.session_state['extracted_data']['normalized_dicts']

            st.success("Successfully extracted financial data!")

            st.subheader("Extracted Financial Data Preview (Top 10 entries)")
            for entry in normalized_dicts[:10]:
                st.json(entry)

            rag_pipeline = init_rag_pipeline(text_chunks)

            question = st.text_input("Ask a question about the financial data:")

            if question:
                answer = find_answer_in_normalized_data_w_yrs(question, normalized_dicts)

                if is_empty_answer(answer):
                    try:
                        rag_result = rag_pipeline.invoke({"question": question})
                        if isinstance(rag_result, dict) and "answer" in rag_result:
                            answer = rag_result["answer"].strip()
                        else:
                            answer = str(rag_result).strip()
                    except Exception as e:
                        st.error(f"Error during RAG processing: {str(e)}")
                        answer = "Sorry, there was a problem processing your request. Please try again later."

                if is_empty_answer(answer):
                    answer = "Sorry, no answer found. Please try rephrasing or upload another document."

                st.markdown(f"**Answer:** {answer}")

        except Exception as e:
            st.error(f"Unexpected error: {str(e)}")
            st.stop()

if not os.path.exists("logs"):
    os.makedirs("logs")

if __name__ == "__main__":
    main()
